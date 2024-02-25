import openpyxl
import subprocess
import time

def run_ollama(text):
    # Run the ollama command and capture the output
    try:
        result = subprocess.run(['ollama', 'run', 'mistral', text.encode("utf-8")], capture_output=True, text=True, timeout=60000)
        return result.stdout.strip()
    except subprocess.TimeoutExpired:
        return "Response timed out"
    except Exception as e:
        return f"Error: {e}"

def main():
    # Load the Excel file
    try:
        wb = openpyxl.load_workbook('Final Processed Dataset.xlsx')
        ws = wb.active
    except FileNotFoundError:
        print("Error: Excel file not found.")
        return
    except Exception as e:
        print(f"Error: {e}")
        return

    # Find the columns for 'Question Title' and 'Question Body'
    title_col_index = None
    body_col_index = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'Question Title':
            title_col_index = col[0].column
        elif col[0].value == 'Question Body':
            body_col_index = col[0].column

    if title_col_index is None or body_col_index is None:
        print("Error: 'Question Title' or 'Question Body' column is not present in the Excel file.")
        return

    # Add a new column for Mistral Response
    response_col_index = ws.max_column + 1
    ws.cell(row=1, column=response_col_index).value = 'Mistral Response'

    # Run the ollama command for each combined text and store the response
    c = 1
    for row in ws.iter_rows(min_row=2, max_row=202):
    # for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        title = row[title_col_index - 1].value or ""
        body = row[body_col_index - 1].value or ""
        combined_text = "Following is a question posted on data science stack exchange, generate a helpful response that is less than 200 words: " + title + " " + body
        # print(type(combined_text))
        # break
        print("Running the model..." , c)
        c+=1
        response = run_ollama(combined_text)
        ws.cell(row=row[0].row, column=response_col_index).value = response
        time.sleep(1)  # Wait for a second before making the next request
        with open("file.txt" , "a") as file:
            file.write(combined_text)
            file.write("\n------\n")
            file.write(response)
            file.write("\n*********\n")
        wb.save('Updated_file_Mistral.xlsx')

    # Save the updated workbook

if __name__ == '__main__':
    main()
